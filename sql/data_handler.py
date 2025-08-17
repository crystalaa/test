import re
import pandas as pd
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook
import xlrd
import gc
import zipfile
import xml.etree.ElementTree as ET

class LoadColumnWorker(QThread):
    """用于在独立线程中读取Excel列名和页签"""
    sheet_names_loaded = pyqtSignal(str, list)  # 发送文件路径和页签列表
    error_occurred = pyqtSignal(str)

    def __init__(self, file_path, sheet_name=None):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name

    def run(self):
        try:
            # 获取所有页签名称
            if self.file_path.lower().endswith('.xlsx'):

                with zipfile.ZipFile(self.file_path, 'r') as zf:
                    # 极少数大文件 workbook.xml 会分片；read() 会一次性读进来，通常 < 200 KB
                    xml_bytes = zf.read('xl/workbook.xml')

                # 去掉命名空间，方便查找
                root = ET.fromstring(xml_bytes)
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                sheet_names = [sheet.attrib['name'] for sheet in root.findall('.//ns:sheet', ns)]
            else:  # .xls
                # with zipfile.ZipFile(self.file_path,) as z:
                #     xml = z.read('xl/workbook.xml')
                # root = ET.fromstring(xml)
                # ns = {'n': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                # sheet_names = [s.attrib['name'] for s in root.findall('.//n:sheet', ns)]
                wb = xlrd.open_workbook(self.file_path)
                sheet_names = wb.sheet_names()
                wb.release_resources()
            self.sheet_names_loaded.emit(self.file_path, sheet_names)
        except Exception as e:
            self.error_occurred.emit(f"读取页签失败: {str(e)}")






def read_excel_fast(file_path, sheet_name, is_file1=True, skip_rows=0, chunk_size=10000):
    """
    快速读取Excel文件，支持大文件分块读取和多表头处理
    优化点：
    1. 分离表头和数据读取，解决read_only模式下无法获取合并单元格的问题
    2. 分块读取大型文件，显著降低内存占用
    3. 及时释放资源，减少内存泄漏
    """
    try:
        if file_path.lower().endswith('.xlsx'):
            # # 阶段1：读取表头和合并单元格信息（使用非只读模式）
            wb = load_workbook(file_path, data_only=True, read_only=False, keep_links=False)
            ws = wb[sheet_name]

            # 读取表头行（最多读取前2行）
            max_header_rows = 2
            header_rows = list(ws.iter_rows(values_only=True, max_row=max_header_rows))
            merged_ranges = list(ws.merged_cells.ranges) if ws.merged_cells else []



            # 处理表头
            if is_file1 and len(header_rows) >= 2 and merged_ranges:
                # 平台文件：处理一级+二级表头
                level1 = [str(v or '') for v in header_rows[0]]
                level2 = [str(v or '') for v in header_rows[1]]

                # 处理一级表头的合并单元格
                for merged in merged_ranges:
                    # merged.bounds格式：(min_row, min_col, max_row, max_col)，索引从1开始
                    if merged.bounds[1] == 1:  # 第1行
                        min_col, max_col = merged.bounds[0], merged.bounds[2]
                        fill_val = level1[min_col - 1]
                        for c in range(min_col, max_col + 1):
                            level1[c - 1] = fill_val

                # 合并两级表头
                cols = [f"{a}-{b}".strip('-') for a, b in zip(level1, level2)]
                data_start_row = 3  # 数据从第3行开始（索引从1开始）

            elif is_file1 and len(header_rows) >= 2 and not merged_ranges:
                # ERP文件或单级表头处理
                header_row_idx = skip_rows
                if len(header_rows) > header_row_idx:
                    cols = [str(v) if v is not None else '' for v in header_rows[header_row_idx]]
                else:
                    cols = []
                data_start_row = header_row_idx + 2  # 数据开始行（索引从1开始）
            elif not is_file1 and not merged_ranges:
                # 非平台文件：处理一级表头
                header_row_idx = skip_rows
                if len(header_rows) > header_row_idx:
                    cols = [str(v) if v is not None else '' for v in header_rows[header_row_idx]]
                else:
                    cols = []
                data_start_row = header_row_idx + 2  # 数据开始行（索引从1开始）
            elif not is_file1 and merged_ranges:

                header_row = 1 if is_file1 else (skip_rows + 2)
                cols = [str(v or '') for v in header_rows[header_row - 1]]
                data_start_row = header_row + 1

            # 清理列名
            cols = [re.sub(r'[\*\s]+', '', c) for c in cols]
            if not cols:
                raise ValueError("未能正确解析表头，请检查文件格式")



            # 获取总数据行数（减去表头行）
            total_rows = ws.max_row
            if total_rows < data_start_row:
                wb.close()
                return pd.DataFrame(columns=cols)  # 空数据框

            # 分块读取数据
            chunks = []
            current_row = data_start_row

            while current_row <= total_rows:
                # 计算当前块的结束行
                end_row = min(current_row + chunk_size - 1, total_rows)

                # 读取当前块数据
                data_rows = list(ws.iter_rows(
                    min_row=current_row,
                    max_row=end_row,
                    values_only=True
                ))

                # 创建数据块DataFrame
                chunk_df = pd.DataFrame(data_rows, columns=cols)
                chunks.append(chunk_df)

                # 更新进度并释放内存
                current_row = end_row + 1
                del data_rows, chunk_df
                gc.collect()

            # 关闭数据工作簿
            wb.close()
            del wb, ws
            gc.collect()

            # 合并所有数据块
            if chunks:
                df = pd.concat(chunks, ignore_index=True)
                del chunks
                gc.collect()
                return df
            else:
                return pd.DataFrame(columns=cols)

        elif file_path.lower().endswith('.xls'):
            max_header_rows = 2
            bk = xlrd.open_workbook(file_path, on_demand=True)
            sh = bk.sheet_by_name(sheet_name)
            header_rows = [
                [str(sh.cell_value(r, c)) if sh.cell_value(r, c) is not None else ''
                 for c in range(sh.ncols)]
                for r in range(min(max_header_rows, sh.nrows))
            ]
            # ---------- 阶段1：读取两级表头 ----------

            level1_raw = [str(sh.cell_value(0, c)).strip()
                          for c in range(sh.ncols)]
            non_empty = sum(1 for v in level1_raw if v)
            empty = sum(1 for v in level1_raw if not v)
            # 真合并标志：只要存在横向合并且覆盖第 0 行即可
            real_merge = any(r1 == 0 and r2 == 0 for r1, r2, _, _ in sh.merged_cells)

            # 视觉合并判定
            visual_merge = (non_empty > 0 and empty > 0) and (not real_merge)

            # ---------- 阶段2：与 xlsx 完全等价的列名生成 ----------
            if is_file1 and (visual_merge or real_merge) :
                # 平台文件：一级+二级表头


                level1 = header_rows[0]
                level2 = header_rows[1]

                # 视觉合并：把左侧非空值向右填充
                last = ''
                for c in range(sh.ncols):
                    if level1[c]:
                        last = level1[c]
                    else:
                        level1[c] = last

                cols = [f"{a}-{b}".strip('-') for a, b in zip(level1, level2)]
                data_start_row = 2  # 行号从 0 开始，数据从第 3 行（索引 2）开始

            elif is_file1 and not visual_merge and not real_merge:
                # ERP文件或单级表头
                header_row_idx = skip_rows + 1
                cols = [str(v) for v in header_rows[header_row_idx]]

                data_start_row = header_row_idx + 1  # 数据行索引（从 0 开始）

            elif not is_file1 and  not visual_merge and not real_merge:
                # 非平台文件：一级表头
                header_row_idx = skip_rows + 1
                cols = [str(v or '') for v in header_rows[header_row_idx]]
                data_start_row = header_row_idx + 1

            elif not is_file1  and (visual_merge or real_merge):
                # 非平台但有合并（罕见）
                header_row_idx = skip_rows + 1
                # cols = [str(v or '') for v in header_rows[header_row_idx]]
                cols = [str(sh.cell_value(skip_rows + 1, c)) for c in range(sh.ncols)]
                data_start_row = header_row_idx + 1
            else:
                # 兜底
                cols = []
                data_start_row = 0

            # 清理列名
            cols = [re.sub(r'[\*\s]+', '', c) for c in cols]

            # 分块读取数据
            chunks = []
            total_rows = sh.nrows
            current_row = data_start_row

            while current_row < total_rows:
                end_row = min(current_row + chunk_size, total_rows)
                data = []
                for r in range(current_row, end_row):
                    row_values = [sh.cell_value(r, c) for c in range(sh.ncols)]
                    data.append(row_values)

                chunk_df = pd.DataFrame(data, columns=cols)
                chunks.append(chunk_df)

                current_row = end_row
                del data, chunk_df
                gc.collect()

            # 释放资源
            bk.release_resources()
            del bk, sh
            gc.collect()

            # 合并数据块
            if chunks:
                df = pd.concat(chunks, ignore_index=True)
                del chunks
                gc.collect()
                return df
            else:
                return pd.DataFrame(columns=cols)

        else:
            raise ValueError(f"不支持的文件格式: {file_path}")

    except Exception as e:
        raise Exception(f"读取Excel文件失败: {str(e)}")

def read_mapping_table(file_path):
    """读取资产分类映射表，返回 DataFrame"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        if '资产分类映射表' not in wb.sheetnames:
            raise Exception("未找到'资产分类映射表'页签")

        ws = wb['资产分类映射表']

        # 只读取第二行作为表头（第二级表头）
        headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=2, max_row=2))]

        # 读取数据行（从第3行开始）
        data = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            data.append(row)

        df = pd.DataFrame(data, columns=headers)
        wb.close()
        return df
    except Exception as e:
        raise Exception(f"读取资产分类映射表时发生错误: {str(e)}")
