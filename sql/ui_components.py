# ui_components.py
import sys
import traceback
import logging
import os
import shutil
import time

from PyQt5.QtWidgets import QWidget, QPushButton, QFileDialog, QLabel, QVBoxLayout, QHBoxLayout, \
    QPlainTextEdit, QTabWidget, QComboBox, QProgressDialog, QApplication
from PyQt5.QtCore import Qt
from openpyxl import load_workbook

from data_handler import LoadColumnWorker
from rule_handler import read_rules
from comparator import CompareWorker
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
import pandas as pd
import polars as pl           # 超大数据用
import xlsxwriter             # 高速写



class ExcelComparer(QWidget):
    """主窗口类"""

    def __init__(self):
        super().__init__()
        self.file1 = ""
        self.file2 = ""
        self.sheet_name1 = ""
        self.sheet_name2 = ""
        self.initUI()
        self.worker = None
        self.summary_data = {}
        self.columns1 = []
        self.columns2 = []
        self.rules = {}  # 存储解析后的规则
        self.rule_file = ""
        # 初始化 worker 变量
        self.worker_sheet1 = None
        self.worker_sheet2 = None
        self.worker_load1 = None
        self.worker_load2 = None
        self.loading_dialog = None
        # 读取规则文件
        self.load_rules_file()

    def load_rules_file(self):
        """加载规则文件"""
        try:
            # 获取exe文件所在目录
            if hasattr(sys, '_MEIPASS'):
                # 打包后的exe环境
                exe_dir = os.path.dirname(sys.executable)
            else:
                # 开发环境
                exe_dir = os.path.dirname(os.path.abspath(__file__))

            rule_file_path = os.path.join(exe_dir, "rule.xlsx")
            self.rule_file = rule_file_path
            if os.path.exists(rule_file_path):
                self.rules = read_rules(rule_file_path)
                self.log(f"✅ 成功加载规则文件: {rule_file_path}")
            else:
                self.log(f"❌ 未找到规则文件: {rule_file_path}")
                # 可以选择是否继续运行或者退出
        except Exception as e:
            self.log(f"❌ 读取规则文件失败: {str(e)}")

    def initUI(self):
        """初始化用户界面"""
        self.setWindowTitle("ERP期初数据核对")
        self.resize(1000, 700)

        main_layout = QVBoxLayout()

        # 文件选择区域
        file_layout = QHBoxLayout()

        left_layout = QVBoxLayout()
        self.label1 = QLabel("未选择平台表")
        self.btn1 = QPushButton("选择平台表")
        self.btn1.clicked.connect(self.select_file1)

        self.sheet_label1 = QLabel("选择平台表页签：")
        self.sheet_combo1 = QComboBox()
        self.sheet_combo1.currentTextChanged.connect(self.on_sheet_selection_changed)

        left_layout.addWidget(self.label1)
        left_layout.addWidget(self.btn1)
        left_layout.addWidget(self.sheet_label1)
        left_layout.addWidget(self.sheet_combo1)

        right_layout = QVBoxLayout()
        self.label2 = QLabel("未选择ERP表")
        self.btn2 = QPushButton("选择ERP表")
        self.btn2.clicked.connect(self.select_file2)

        self.sheet_label2 = QLabel("选择ERP表页签：")
        self.sheet_combo2 = QComboBox()
        self.sheet_combo2.currentTextChanged.connect(self.on_sheet_selection_changed)

        right_layout.addWidget(self.label2)
        right_layout.addWidget(self.btn2)
        right_layout.addWidget(self.sheet_label2)
        right_layout.addWidget(self.sheet_combo2)
        file_layout.addLayout(left_layout)
        file_layout.addLayout(right_layout)
        # 按钮区域
        button_layout = QHBoxLayout()
        self.compare_btn = QPushButton("比较文件")
        self.compare_btn.setFixedWidth(150)
        self.compare_btn.clicked.connect(self.compare_files)
        self.compare_btn.setEnabled(False)
        self.export_btn = QPushButton("导出报告")
        self.export_btn.setFixedWidth(150)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_report)
        button_layout.addStretch()
        button_layout.addWidget(self.compare_btn)
        button_layout.addWidget(self.export_btn)
        # 日志和报告区域
        self.tab_widget = QTabWidget()
        self.log_area = QPlainTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setStyleSheet("background-color: #f0f0f0;")
        self.summary_area = QPlainTextEdit()
        self.summary_area.setReadOnly(True)
        self.summary_area.setStyleSheet("background-color: #f0f0f0;")
        self.tab_widget.addTab(self.log_area, "比对日志")
        self.tab_widget.addTab(self.summary_area, "汇总报告")
        # 主布局组合
        main_layout.addLayout(file_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.tab_widget)

        self.setLayout(main_layout)

    def closeEvent(self, event):
        """窗口关闭时确保线程安全退出"""
        if hasattr(self, 'worker') and self.worker is not None and self.worker.isRunning():
            self.worker.quit()
            self.worker.wait()
        if hasattr(self, 'worker_load1') and self.worker_load1 is not None and self.worker_load1.isRunning():
            self.worker_load1.quit()
            self.worker_load1.wait()
        if hasattr(self, 'worker_load2') and self.worker_load2 is not None and self.worker_load2.isRunning():
            self.worker_load2.quit()
            self.worker_load2.wait()
        if hasattr(self, 'worker_sheet1') and self.worker_sheet1 is not None and self.worker_sheet1.isRunning():
            self.worker_sheet1.quit()
            self.worker_sheet1.wait()
        if hasattr(self, 'worker_sheet2') and self.worker_sheet2 is not None and self.worker_sheet2.isRunning():
            self.worker_sheet2.quit()
            self.worker_sheet2.wait()
        super().closeEvent(event)

    def reset_file_state(self, is_file1=True, is_file2=False):
        if is_file1:
            self.columns1 = []
            self.sheet_combo1.clear()
            self.sheet_combo1.setEnabled(True)
            self.sheet_label1.setText("选择平台表页签：")
            if hasattr(self, 'worker_sheet1'):
                self.worker_sheet1 = None
        if is_file2:
            self.columns2 = []
            self.sheet_combo2.clear()
            self.sheet_combo2.setEnabled(True)
            self.sheet_label2.setText("选择ERP表页签：")
            if hasattr(self, 'worker_sheet2'):
                self.worker_sheet2 = None
        self.compare_btn.setEnabled(False)
        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

    def select_file1(self):
        self.reset_file_state(is_file1=True, is_file2=False)
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file1 = file
            filename = os.path.basename(file)
            self.label1.setText(f"平台表: {filename}")
            # 显示加载对话框
            self.show_loading_dialog("正在加载平台表页签...")
            self.load_sheet_and_columns(file, is_file1=True)

    def select_file2(self):
        self.reset_file_state(is_file1=False, is_file2=True)
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file2 = file
            filename = os.path.basename(file)

            self.label2.setText(f"ERP表: {filename}")
            self.show_loading_dialog("正在加载ERP表页签...")
            self.load_sheet_and_columns(file, is_file2=True)

    def show_loading_dialog(self, message="正在加载，请稍候..."):
        """显示加载对话框"""
        if not self.loading_dialog:
            self.loading_dialog = QProgressDialog(message, None, 0, 0, self)
            self.loading_dialog.setWindowModality(Qt.WindowModal)
            self.loading_dialog.setWindowTitle("加载中")
            self.loading_dialog.setCancelButton(None)
            self.loading_dialog.show()

    def load_sheet_and_columns(self, file_path, is_file1=False, is_file2=False):

        worker = LoadColumnWorker(file_path)
        worker.sheet_names_loaded.connect(self.on_sheet_names_loaded)
        worker.sheet_names_loaded.connect(self.close_loading_dialog)
        # worker.columns_loaded.connect(self.on_columns_loaded)
        # worker.error_occurred.connect(self.on_column_error)
        if is_file1:
            self.worker_load1 = worker
        elif is_file2:
            self.worker_load2 = worker
        worker.start()

    def on_sheet_names_loaded(self, file_path, sheet_names):
        if file_path == self.file1:
            self.sheet_combo1.clear()
            self.sheet_combo1.addItems(sheet_names)
            self.sheet_combo1.setCurrentIndex(0)
        elif file_path == self.file2:
            self.sheet_combo2.clear()
            self.sheet_combo2.addItems(sheet_names)
            self.sheet_combo2.setCurrentIndex(0)

    def on_sheet_selection_changed(self):
        """页签选择变化时的处理函数"""
        # 简单更新比较按钮状态
        self.update_compare_button_state()

    def update_compare_button_state(self):
        sheet_selected = self.sheet_combo1.currentText() and self.sheet_combo2.currentText()
        if not sheet_selected:
            self.compare_btn.setEnabled(False)
            return

        self.compare_btn.setEnabled(True)

    def compare_files(self):
        if not self.file1 or not self.file2:
            self.log("请先选择两个 Excel 文件！")
            return
        sheet_name1 = self.sheet_combo1.currentText()
        sheet_name2 = self.sheet_combo2.currentText()
        if not sheet_name1 or not sheet_name2:
            self.log("请选择两个文件的页签！")
            return

        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

        # 获取主键字段
        primary_keys = [field for field, rule in self.rules.items() if rule["is_primary"]]
        if not primary_keys:
            self.log("规则文件中未定义主键字段，请检查规则文件！")
            return
        self.loading_dialog = QProgressDialog("正在比较文件，请稍候...", None, 0, 0, self)
        self.loading_dialog.setWindowModality(Qt.WindowModal)
        self.loading_dialog.setWindowTitle("比较中")
        self.loading_dialog.setCancelButton(None)
        self.loading_dialog.show()

        self.worker = CompareWorker(self.file1, self.file2, self.rule_file, sheet_name1, sheet_name2,
                                    primary_keys=primary_keys,
                                    rules=self.rules)
        self.worker.log_signal.connect(self.log)
        # 连接信号以在比较完成时关闭对话框
        self.worker.finished.connect(self.close_loading_dialog)
        self.worker.finished.connect(lambda: self.export_btn.setEnabled(True))
        self.worker.finished.connect(self.on_compare_finished)
        self.worker.start()

    def close_loading_dialog(self):
        """关闭加载对话框"""
        if self.loading_dialog:
            self.loading_dialog.close()
            self.loading_dialog = None

    def on_compare_finished(self):
        try:
            if hasattr(self.worker, 'summary'):
                self.summary_data = self.worker.summary
                primary_key = self.summary_data.get("primary_key", "主键")
                total_file1 = self.summary_data['total_file1']
                total_file2 = self.summary_data['total_file2']
                missing_count = self.summary_data['missing_count']
                extra_count = self.summary_data.get('extra_count', 0)
                common_count = self.summary_data['common_count']
                diff_count = self.summary_data['diff_count']
                equal_count = self.summary_data['equal_count']
                diff_ratio = self.summary_data['diff_ratio']
                missing_columns = self.summary_data.get("missing_columns", [])
                missing_columns_str = ", ".join(missing_columns) if missing_columns else "无"

                summary_text = (
                    f"📊 比对汇总报告\n"
                    f"--------------------------------\n"
                    f"• 总{primary_key}数量（平台表）：{total_file1}\n"
                    f"• 总{primary_key}数量（ERP表）：{total_file2}\n"
                    f"• ERP表中缺失的{primary_key}：{missing_count}\n"
                    f"• ERP表中多出的{primary_key}：{extra_count}\n"
                    f"• 共同{primary_key}数量：{common_count}\n"
                    f"• 列不一致的{primary_key}数量：{diff_count}\n"
                    f"• 列一致的{primary_key}数量：{equal_count}\n"
                    f"• ERP表中缺失的列：{missing_columns_str}\n"
                    f"--------------------------------\n"
                    f"• 差异数据占比：{diff_ratio:.2%}\n"
                )
                self.summary_area.setPlainText(summary_text)
                self.export_btn.setEnabled(True)
        except Exception as e:
            self.summary_area.setPlainText(f"❌ 显示汇总报告时发生错误：{str(e)}\n请查看比对日志了解详细信息。")
            self.export_btn.setEnabled(False)

    # ---------- 导出入口 ----------
    def export_report(self):
        if not hasattr(self.worker, 'diff_full_rows'):
            self.log("没有可导出的数据，请先执行比对！")
            return
        directory = QFileDialog.getExistingDirectory(self, "选择保存路径")
        if not directory:
            return
        tasks = [
            (self.file1, self.sheet_combo1.currentText(), True, directory),
            (self.file2, self.sheet_combo2.currentText(), False, directory)
        ]
        t0 = time.time()
        self.loading_dialog = QProgressDialog("正在导出报告，请稍候...", None, 0, 0, self)
        self.loading_dialog.setWindowModality(Qt.WindowModal)
        self.loading_dialog.setWindowTitle("导出")
        self.loading_dialog.setCancelButton(None)
        self.loading_dialog.show()
        with ThreadPoolExecutor(max_workers=2) as pool:
            pool.map(lambda t: self._export_final(*t), tasks)
        self.log(f"✅ 并行导出完成，总耗时 {time.time() - t0:.1f}s")
        self.close_loading_dialog()

    # ---------- 最终导出实现 ----------
    def _export_final(self, src_file, sheet_name, is_first_file, out_dir):
        try:

            # 1. 复制原文件
            dst = Path(out_dir) / f"{Path(src_file).stem}_比对结果.xlsx"
            shutil.copy2(src_file, dst)
            # 1.1 去掉只读属性（Windows / Linux / macOS 通用）
            try:
                os.chmod(dst, 0o666)  # Linux / macOS
            except Exception:
                pass  # Windows 会抛异常，忽略即可
            wb = load_workbook(filename=dst, read_only=False)
            ws = wb[sheet_name]
            has_merged_cell = False
            for row in ws.iter_rows(max_row=2):
                for cell in row:
                    if cell.coordinate in ws.merged_cells:
                        has_merged_cell = True
                        break
            if not is_first_file:
                # 遍历合并单元格范围
                if has_merged_cell:
                    df = pd.read_excel(dst, sheet_name=sheet_name, skiprows=1, dtype=str).fillna("")
                else:
                    # 2. 读原表（全部字符串，防类型问题）
                    df = pd.read_excel(dst, sheet_name=sheet_name, dtype=str).fillna("")
            else:
                df = pd.read_excel(dst, sheet_name=sheet_name, dtype=str).fillna("")

            # 3. 动态主键字段
            primary_keys = [f for f, r in self.rules.items() if r.get("is_primary")]

            # 4. 计算行主键（与比对阶段一致）
            if is_first_file:
                # 平台表：直接取主键列
                df["_key"] = df[primary_keys].astype(str).agg(" + ".join, axis=1)
            else:
                # ERP表：根据规则里的计算表达式动态生成
                pk_field = next(f for f, r in self.rules.items() if r.get("is_primary"))
                rule = self.rules[pk_field]
                if rule.get("calc_rule"):
                    df["_key"] = self.worker.calculate_field(df, rule["calc_rule"], rule["data_type"]).astype(str)
                else:
                    df["_key"] = df[rule["table2_field"]].astype(str)

            # 5. 建立差异映射
            diff_map, miss, extra = {}, set(), set()
            for it in getattr(self.worker, 'diff_full_rows', []):
                key = " + ".join([str(it['source' if is_first_file else 'target'].get(pk, ""))
                                  for pk in primary_keys])
                diff_map[key] = it
            for row in getattr(self.worker, 'missing_rows', []):
                miss.add(" + ".join([str(row.get(pk, "")) for pk in primary_keys]))
            for row in getattr(self.worker, 'extra_in_file2', []):
                extra.add(" + ".join([str(row.get(pk, "")) for pk in primary_keys]))

            # 6. 需要追加的列（顺序 = 规则顺序）
            comp_cols = [f for f in self.rules.keys() if not self.rules[f].get("is_primary")]

            # 7. 计算追加值
            keys = df["_key"].tolist()
            comp_results = [
                "此数据不存在于SAP" if k in miss else  # 平台表多余 → 提示不存在于SAP
                "此数据不存在于平台" if k in extra else  # ERP表多余 → 提示不存在于平台
                "不一致" if k in diff_map else
                "一致"
                for k in keys
            ]

            def detail(row_key, fld):
                if row_key not in diff_map:
                    return ""
                s, t = diff_map[row_key]['source'], diff_map[row_key]['target']

                # 资产分类特殊处理：用中文提示
                if fld == "资产分类":
                    code1 = s.get(fld, "")
                    code2 = t.get('原21版资产分类', "")
                    v1 = self.worker.asset_code_map.get(code1, code1)
                    v2 = t.get(fld)

                    # 资产分类使用code1和code2进行比较
                    if self.worker.values_equal_by_rule(code1, code2,
                                                        self.rules[fld]["data_type"],
                                                        self.rules[fld].get("tail_diff"),
                                                        fld):
                        return ""
                    return f"不一致：平台表={v1 or ''}, ERP表={v2 or ''}"
                else:
                    # 其他字段使用v1和v2进行比较
                    v1 = self.normalize_value(s.get(fld, ""))
                    v2 = self.normalize_value(t.get(fld, ""))

                    if self.worker.values_equal_by_rule(v1, v2,
                                                        self.rules[fld]["data_type"],
                                                        self.rules[fld].get("tail_diff"),
                                                        fld):
                        return ""
                    return f"不一致：平台表={v1 or ''}, ERP表={v2 or ''}"

            comp_details = {
                fld: [detail(k, fld) for k in keys]
                for fld in comp_cols
            }

            # 8. 用 xlsxwriter 重写副本：不改动原列，仅追加
            with xlsxwriter.Workbook(dst, {'nan_inf_to_errors': True}) as wb:
                ws = wb.add_worksheet(sheet_name)
                header_fmt = wb.add_format({'bold': True, 'bg_color': '#FFC7CE'})
                red_fmt = wb.add_format({'bg_color': '#FF0000', 'font_color': '#FFFFFF'})

                orig_cols = len(df.columns) - 1  # 去掉 _key
                orig_rows = len(df)

                # 原标题
                for c, col_name in enumerate(df.columns[:-1]):
                    ws.write(0, c, col_name, header_fmt)
                # 原数据
                for r in range(orig_rows):
                    for c in range(orig_cols):
                        ws.write(r + 1, c, df.iloc[r, c])

                # 追加“对比结果”
                next_col = orig_cols
                ws.write(0, next_col, "对比结果", header_fmt)
                for r in range(orig_rows):
                    val = comp_results[r]
                    ws.write(r + 1, next_col, val)
                    if val != "一致":
                        ws.write(r + 1, next_col, val, red_fmt)

                # 依次追加规则字段列
                for fld in comp_cols:
                    next_col += 1
                    ws.write(0, next_col, fld, header_fmt)
                    for r in range(orig_rows):
                        val = comp_details[fld][r]
                        ws.write(r + 1, next_col, val)
                        if val:
                            ws.write(r + 1, next_col, val, red_fmt)

            self.log(f"✅ 导出完成 {dst.name}")
        except Exception as e:
            self.log(f"❌ 导出失败 {Path(src_file).name}: {e}")

    def _rename_erp_columns(self, df, rules):
        """
        把 ERP 的 Unnamed: X 列名，按规则顺序映射成 table2_field，
        使得 calculate_field 里的字段名都能匹配到真实列。
        """
        # 建立“规则顺序 -> 实际列名”映射
        rename_map = {}
        for rule_field, rule in rules.items():
            tbl2 = rule["table2_field"]
            # 找到实际列名（按顺序）
            if tbl2 in df.columns:
                # 已经对齐，无需改名
                continue
            # 如果规则写的是“公司代码”，但列名是 Unnamed: 1，则手动映射
            # 这里采用“位置映射”：规则顺序与实际列顺序一致
            # 例如：规则第 1 个 table2_field -> df 第 1 列
            # 需要用户保证顺序一致；若不一致，可在规则里加“顺序号”字段
            idx = list(rules.keys()).index(rule_field)
            if idx < len(df.columns):
                rename_map[df.columns[idx]] = tbl2
        return df.rename(columns=rename_map)
    # ---------- 单文件导出 ----------
    def _export_one_file(self, src_file, sheet_name, is_first_file, out_dir):
        try:
            # 自动选引擎：行数>5万→polars；否则pandas+xlsxwriter
            quick_count = self._quick_row_count(src_file, sheet_name)
            use_polars = quick_count > 50_000
            dst = Path(out_dir) / f"{Path(src_file).stem}_比对结果.xlsx"
            if use_polars:
                self._write_with_polars(src_file, sheet_name, is_first_file, dst)
            else:
                self._write_with_xlsxwriter(src_file, sheet_name, is_first_file, dst)
        except Exception as e:
            self.log(f"❌ 导出失败 {Path(src_file).name}: {e}")

    # ---------- 快速估算行数 ----------
    def _quick_row_count(self, file_path, sheet_name):
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
            with pd.ExcelFile(file_path) as xls:
                return xls.book.sheet_by_name(sheet_name).nrows
        except:
            return 0

    # ---------- 方案A：xlsxwriter + pandas ----------
    def _write_with_xlsxwriter(self, src_file, sheet_name, is_first_file, dst_file):
        # 1) 读数据
        df = pd.read_excel(src_file, sheet_name=sheet_name)

        # 2) 清理 NaN/Inf
        df = df.replace([float('inf'), float('-inf')], None)  # 先转 None
        df = df.where(pd.notnull(df), None)  # 再转 None（覆盖 NaN）

        # 3) 计算对比列
        df = self._add_comparison_columns(df, is_first_file)

        # 4) 写文件，打开 nan_inf 容错
        with xlsxwriter.Workbook(
                dst_file,
                {
                    'constant_memory': True,
                    'nan_inf_to_errors': True  # ← 关键
                }
        ) as wb:
            ws = wb.add_worksheet(sheet_name)

            header_fmt = wb.add_format({'bold': True, 'bg_color': '#FFC7CE'})
            red_fmt = wb.add_format({'bg_color': '#FF0000', 'font_color': '#FFFFFF'})

            # 写标题
            for c, col in enumerate(df.columns):
                ws.write(0, c, col, header_fmt)

            # 批量写数据（None 会自动写成空单元格）
            for r, row in enumerate(df.itertuples(index=False), start=1):
                for c, val in enumerate(row):
                    ws.write(r, c, "" if val is None else val)

            # 标记差异行
            try:
                res_idx = df.columns.get_loc("对比结果")
                for r in range(1, len(df) + 1):
                    val = df.iloc[r - 1, res_idx]
                    if val != "一致":
                        ws.write(r, res_idx, val or "", red_fmt)
            except Exception:
                pass

        self.log(f"✅ xlsxwriter 导出完成 {Path(dst_file).name}")

    # ---------- 方案B：polars 零拷贝 ----------
    def _write_with_polars(self, src_file, sheet_name, is_first_file, dst_file):
        # 1) 读为 polars DataFrame
        df = pl.read_excel(src_file, sheet_name=sheet_name)
        # 2) 计算对比列（复用逻辑，转回 pandas 计算后转回 polars，极快）
        pdf = self._add_comparison_columns(df.to_pandas(), is_first_file)
        # 3) 写
        pl.from_pandas(pdf).write_excel(dst_file, worksheet=sheet_name)
        self.log(f"✅ polars 导出完成 {dst_file.name}")

    # ---------- 计算对比列（复用原逻辑，稍作适配） ----------
    def _add_comparison_columns(self, df: pd.DataFrame, is_first_file: bool):
        primary_keys = [f for f, r in self.rules.items() if r["is_primary"]]
        compare_cols = list(self.rules.keys())

        df = df.copy()
        # 主键列
        df["_key"] = df[primary_keys].astype(str).agg(" + ".join, axis=1)

        # 差异映射
        diff_map, miss, extra = {}, set(), set()
        for it in getattr(self.worker, 'diff_full_rows', []):
            key = " + ".join([str(it['source' if is_first_file else 'target'].get(pk, '')) for pk in primary_keys])
            diff_map[key] = it
        for row in getattr(self.worker, 'missing_rows', []):
            miss.add(" + ".join([str(row.get(pk, '')) for pk in primary_keys]))
        for row in getattr(self.worker, 'extra_in_file2', []):
            extra.add(" + ".join([str(row.get(pk, '')) for pk in primary_keys]))

        # 对比结果
        def comp(row):
            k = row["_key"]
            if k in miss:
                return "此数据不存在于SAP" if is_first_file else "此数据不存在于平台"
            if k in extra:
                return "此数据不存在于平台" if is_first_file else "此数据不存在于SAP"
            return "不一致" if k in diff_map else "一致"

        df["对比结果"] = df.apply(comp, axis=1)

        # 各列差异详情
        for col in compare_cols:
            if col not in df.columns:
                continue

            def detail(row):
                k = row["_key"]
                if k not in diff_map:
                    return ""
                s, t = diff_map[k]['source'], diff_map[k]['target']
                v1, v2 = s.get(col, ""), t.get(col, "")
                rule = self.rules.get(col, {})
                if not self.worker.values_equal_by_rule(v1, v2, rule.get("data_type"), rule.get("tail_diff"), col):
                    return f"不一致：平台表={v1}, ERP表={v2}"
                return ""

            df[col] = df.apply(detail, axis=1)

        return df.drop(columns=["_key"])
    def log(self, message):
        """日志输出"""
        self.log_area.appendPlainText(message)


    @staticmethod
    def normalize_value(val):
        """统一空值表示"""
        if pd.isna(val) or val is None or (isinstance(val, str) and str(val).strip() == ''):
            return ''
        return str(val).strip()


def exception_hook(exc_type, exc_value, exc_traceback):
    """全局异常钩子，防止崩溃"""
    try:
        ex = QApplication.instance().topLevelWidgets()[0]
        if hasattr(ex, "log"):
            error_message = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
            logging.error(error_message)
            ex.log(f"❌ 发生异常：{exc_value}")
        else:
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
    except:
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
